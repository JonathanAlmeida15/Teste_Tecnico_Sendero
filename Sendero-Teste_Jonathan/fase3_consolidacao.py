"""
FASE 3 — CONSOLIDAÇÃO E CLASSIFICAÇÃO POR REGRAS
Aplica regras baseadas em palavras-chave ao texto higienizado,
calcula scores (pesos base + bônus por combinação) e gera
a planilha final .xlsx com três abas.
"""

import os
import pathlib
import re
import logging
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Diretório base (pasta onde está este .py) ─────────────────────────────────
BASE_DIR = pathlib.Path(__file__).resolve().parent

ARQUIVO_SAIDA = BASE_DIR / "pipeline_pje_output.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# 1. REGRAS DE CLASSIFICAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

# Cada tema tem um peso base e uma lista de padrões (regex) para detectá-lo.
# Os padrões já operam sobre texto normalizado (sem acento, minúsculas).

TEMAS: dict[str, dict] = {
    "homologacao": {
        "peso": 30,
        "padroes": [
            r"\bhomolog",                           # homologação, homologado, homologa…
            r"\baprovac[a-z]*\s+judicial",
            r"\baprovado\s+pelo\s+juiz",
        ],
    },
    "rateio_pagamento": {
        "peso": 30,
        "padroes": [
            r"\brateio",
            r"\bpagamento",
            r"\bpagamentos",
            r"\bdistribuic[a-z]*\s+de\s+valores",
            r"\bpagos\s+aos\s+credores",
        ],
    },
    "credor_silente": {
        "peso": 25,
        "padroes": [
            r"\bcredor\s+silente",
            r"\bcredores\s+silentes",
            r"\bsilente",
            r"\bnao\s+(se\s+)?manifest",           # não se manifestou
            r"\bausenica\s+de\s+manifestac",
        ],
    },
    "conta_judicial": {
        "peso": 20,
        "padroes": [
            r"\bconta\s+judicial",
            r"\bconta\s+unificada",
            r"\bsaldo\s+judicial",
            r"\bsaldo\b",
            r"\bdeposito\s+judicial",
            r"\bvalores?\s+depositados?",
        ],
    },
    "prazo": {
        "peso": 18,
        "padroes": [
            r"\bprazo",
            r"\bprazos",
            r"\bdias?\s+uteis?",
            r"\bdias?\s+corridos?",
            r"\bvencimento\s+do\s+prazo",
            r"\bintimac[a-z]*\s+para\s+cumpr",
        ],
    },
    "decisao": {
        "peso": 15,
        "padroes": [
            r"\bdecisao",
            r"\bdecidiu",
            r"\bdecide",
            r"\bsentenc",                           # sentença, sentenciado…
            r"\bacordao",
            r"\bjulgamento",
            r"\bjulgou",
        ],
    },
    "edital": {
        "peso": 12,
        "padroes": [
            r"\bedital",
            r"\bpublicac[a-z]*\s+oficial",
            r"\bdiario\s+oficial",
            r"\bpublicado\s+em\s+edital",
        ],
    },
    "cessao_credito": {
        "peso": 12,
        "padroes": [
            r"\bcess[a-z]*\s+de\s+credito",
            r"\bcessiona",
            r"\bcess[a-z]*\s+creditorias?",
            r"\btransferencia\s+de\s+credito",
            r"\bcedente",
            r"\bcessionario",
        ],
    },
    "despacho": {
        "peso": 8,
        "padroes": [
            r"\bdespacho",
            r"\bdespachos",
            r"\bdespachado",
        ],
    },
    "peticao": {
        "peso": 5,
        "padroes": [
            r"\bpetic[a-z]*",                       # petição, peticionou…
            r"\bmanifestac[a-z]*\s+das?\s+partes?",
            r"\brequerimento",
        ],
    },
}

# Termos auxiliares para bônus (não são temas autônomos)
PADROES_QGC      = [r"\bqgc\b", r"\bquadro\s+geral\s+de\s+credores?"]
PADROES_CREDORES = [r"\bcredores?\b", r"\bcredora\b"]
PADROES_PERDA    = [r"\bperda\s+de\s+direito", r"\bperda\s+de\s+valor", r"\bprescric"]

# Combinações de bônus: (tema_a, tema_b_ou_auxiliar, bonus, descrição)
BONUS_COMBINACOES: list[tuple] = [
    ("homologacao",      "rateio_pagamento", 40, "Homologação + Rateio/Pagamento"),
    ("homologacao",      "__qgc__",          40, "Homologação + QGC"),
    ("decisao",          "prazo",            15, "Decisão + Prazo"),
    ("conta_judicial",   "__saldo__",        10, "Conta judicial + Saldo"),
    ("edital",           "__credores__",     10, "Edital + Credores"),
    ("credor_silente",   "prazo",            15, "Credor silente + Prazo"),
    ("credor_silente",   "__perda__",        15, "Credor silente + Perda de valor"),
]

# Auxiliares especiais usados em bônus
AUXILIARES: dict[str, list[str]] = {
    "__qgc__":      PADROES_QGC,
    "__saldo__":    [r"\bsaldo\b"],
    "__credores__": PADROES_CREDORES,
    "__perda__":    PADROES_PERDA,
}


# ══════════════════════════════════════════════════════════════════════════════
# 2. FUNÇÕES DE CLASSIFICAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def normalizar_para_busca(texto: str) -> str:
    """
    Remove acentos e converte para minúsculas — usado apenas internamente
    para casamento de padrões, não altera o texto que vai para a planilha.
    """
    if not texto or not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize("NFD", texto.lower())
    texto = re.sub(r"[\u0300-\u036f]", "", texto)
    return texto


def detectar_tema(texto_norm: str, padroes: list[str]) -> tuple[bool, list[str]]:
    """Retorna (encontrado, lista de trechos que dispararam o tema)."""
    evidencias = []
    for padrao in padroes:
        for match in re.finditer(padrao, texto_norm):
            start = max(0, match.start() - 20)
            end   = min(len(texto_norm), match.end() + 20)
            trecho = "…" + texto_norm[start:end].strip() + "…"
            evidencias.append(trecho)
    return bool(evidencias), evidencias


def classificar_comunicacao(texto_higienizado: str) -> dict:
    """
    Recebe o texto higienizado e retorna um dicionário com:
      - temas:       lista de temas detectados
      - score:       pontuação final
      - bonus_desc:  bônus aplicados
      - evidencias:  trechos que ativaram cada tema
    """
    texto_norm = normalizar_para_busca(texto_higienizado)

    temas_detectados: dict[str, list[str]] = {}

    for nome_tema, cfg in TEMAS.items():
        encontrado, evidencias = detectar_tema(texto_norm, cfg["padroes"])
        if encontrado:
            temas_detectados[nome_tema] = evidencias

    # Score base
    score_base = sum(TEMAS[t]["peso"] for t in temas_detectados)

    # Bônus por combinação
    bonus_total = 0
    bonus_aplicados: list[str] = []

    for (tema_a, tema_b, bonus, descricao) in BONUS_COMBINACOES:
        # Verifica tema_a (sempre um tema real)
        tem_a = tema_a in temas_detectados

        # Verifica tema_b (pode ser tema real ou auxiliar)
        if tema_b.startswith("__"):
            padroes_b = AUXILIARES[tema_b]
            tem_b, _ = detectar_tema(texto_norm, padroes_b)
        else:
            tem_b = tema_b in temas_detectados

        if tem_a and tem_b:
            bonus_total += bonus
            bonus_aplicados.append(f"{descricao} (+{bonus})")

    score_final = score_base + bonus_total

    # Formata evidências como string legível
    evidencias_str = "; ".join(
        f"[{t.upper()}]: {', '.join(ev[:2])}"          # até 2 trechos por tema
        for t, ev in temas_detectados.items()
    )

    return {
        "temas":       list(temas_detectados.keys()),
        "score":       score_final,
        "bonus_desc":  "; ".join(bonus_aplicados) if bonus_aplicados else "",
        "evidencias":  evidencias_str,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 3. CONSOLIDAÇÃO POR PROCESSO
# ══════════════════════════════════════════════════════════════════════════════

def consolidar_por_processo(df_comuns: pd.DataFrame) -> pd.DataFrame:
    """Gera visão resumida por processo."""
    rows = []

    for processo, grupo in df_comuns.groupby("numero_processo", sort=False):
        score_total = grupo["score"].sum()
        score_max   = grupo["score"].max()

        # Última comunicação relevante (score > 0 ou simplesmente a mais recente)
        relevantes = grupo[grupo["score"] > 0].copy()
        df_ref = relevantes if not relevantes.empty else grupo
        ultima_data = (
            df_ref["data_comunicacao"]
            .dropna()
            .sort_values(ascending=False)
            .iloc[0]
            if df_ref["data_comunicacao"].notna().any()
            else None
        )

        # Todos os temas distintos do processo
        todos_temas: list[str] = []
        for temas_lista in grupo["temas"].dropna():
            if isinstance(temas_lista, list):
                todos_temas.extend(temas_lista)
            elif isinstance(temas_lista, str) and temas_lista:
                todos_temas.extend(temas_lista.split("; "))
        temas_unicos = list(dict.fromkeys(todos_temas))   # preserva ordem

        # Observação automática
        if score_max >= 100:
            obs = "ALTA RELEVÂNCIA — requer atenção imediata."
        elif score_max >= 50:
            obs = "Relevância moderada — acompanhamento recomendado."
        elif score_max > 0:
            obs = "Atividade rotineira identificada."
        else:
            obs = "Sem temas relevantes identificados no período."

        rows.append({
            "numero_processo":         processo,
            "total_comunicacoes":      len(grupo),
            "score_total":             score_total,
            "score_maximo":            score_max,
            "principais_temas":        "; ".join(temas_unicos),
            "data_ultima_comunicacao": ultima_data,
            "observacao":              obs,
        })

    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# 4. GERAÇÃO DA PLANILHA XLSX
# ══════════════════════════════════════════════════════════════════════════════

# Paleta de cores
COR_CABECALHO  = "1F3864"   # azul escuro
COR_HEADER_TXT = "FFFFFF"
COR_LINHA_PAR  = "DCE6F1"   # azul muito claro
COR_LINHA_IMP  = "FFFFFF"
COR_ALERTA_HDR = "C00000"   # vermelho escuro
COR_RESUMO_HDR = "375623"   # verde escuro


def _estilo_cabecalho(ws, linha: int, fill_hex: str) -> None:
    fill = PatternFill("solid", fgColor=fill_hex)
    fonte = Font(bold=True, color=COR_HEADER_TXT, size=11)
    alin  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[linha]:
        cell.fill    = fill
        cell.font    = fonte
        cell.alignment = alin


def _autofit(ws, max_largura: int = 80) -> None:
    """Ajusta a largura das colunas com base no conteúdo."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_largura)


def _zebra(ws, inicio_linha: int) -> None:
    """Aplica padrão zebra a partir da linha informada."""
    fill_par = PatternFill("solid", fgColor=COR_LINHA_PAR)
    fill_imp = PatternFill("solid", fgColor=COR_LINHA_IMP)
    alin_wrap = Alignment(wrap_text=True, vertical="top")
    for i, row in enumerate(ws.iter_rows(min_row=inicio_linha), start=1):
        fill = fill_par if i % 2 == 0 else fill_imp
        for cell in row:
            cell.fill      = fill
            cell.alignment = alin_wrap


def gerar_planilha(
    df_comuns:       pd.DataFrame,
    df_resumo:       pd.DataFrame,
    alertas_total:   list[dict],
    caminho:         str = ARQUIVO_SAIDA,
) -> None:
    """Gera o arquivo .xlsx com as três abas exigidas."""

    # ── Aba 1: Comunicações classificadas ────────────────────────────────────
    colunas_aba1 = {
        "numero_processo":   "Número do Processo",
        "data_comunicacao":  "Data da Comunicação",
        "tipo_comunicacao":  "Tipo",
        "texto_higienizado": "Texto Higienizado",
        "temas_str":         "Temas Identificados",
        "score":             "Score",
        "bonus_desc":        "Bônus Aplicados",
        "evidencias":        "Evidências / Trechos",
    }

    df_aba1 = df_comuns.copy()
    df_aba1["temas_str"] = df_aba1["temas"].apply(
        lambda t: "; ".join(t) if isinstance(t, list) else (t or "")
    )
    df_aba1 = df_aba1.rename(columns=colunas_aba1)[list(colunas_aba1.values())]
    df_aba1 = df_aba1.sort_values("Score", ascending=False)

    # ── Aba 2: Resumo por processo ────────────────────────────────────────────
    colunas_aba2 = {
        "numero_processo":         "Número do Processo",
        "total_comunicacoes":      "Total de Comunicações",
        "score_total":             "Score Total",
        "score_maximo":            "Score Máximo",
        "principais_temas":        "Principais Temas",
        "data_ultima_comunicacao": "Data Última Comunicação",
        "observacao":              "Observação",
    }
    df_aba2 = df_resumo.rename(columns=colunas_aba2)[list(colunas_aba2.values())]
    df_aba2 = df_aba2.sort_values("Score Total", ascending=False)

    # ── Aba 3: Inconsistências / Alertas ─────────────────────────────────────
    df_aba3 = pd.DataFrame(alertas_total) if alertas_total else pd.DataFrame(
        columns=["processo", "tipo", "descricao"]
    )

    # ── Escreve no xlsx ───────────────────────────────────────────────────────
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df_aba1.to_excel(writer, sheet_name="Comunicações Classificadas", index=False)
        df_aba2.to_excel(writer, sheet_name="Resumo por Processo",        index=False)
        df_aba3.to_excel(writer, sheet_name="Inconsistências e Alertas",  index=False)

    # ── Aplica estilos ────────────────────────────────────────────────────────
    wb = load_workbook(caminho)

    ws1 = wb["Comunicações Classificadas"]
    _estilo_cabecalho(ws1, 1, COR_CABECALHO)
    _zebra(ws1, 2)
    _autofit(ws1)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    ws2 = wb["Resumo por Processo"]
    _estilo_cabecalho(ws2, 1, COR_RESUMO_HDR)
    _zebra(ws2, 2)
    _autofit(ws2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    ws3 = wb["Inconsistências e Alertas"]
    _estilo_cabecalho(ws3, 1, COR_ALERTA_HDR)
    _zebra(ws3, 2)
    _autofit(ws3)
    ws3.freeze_panes = "A2"

    wb.save(caminho)
    logger.info(f"Planilha salva em '{caminho}'.")


# ══════════════════════════════════════════════════════════════════════════════
# 5. FUNÇÃO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def executar(
    df_higienizado: pd.DataFrame | None = None,
    alertas_anteriores: list[dict] | None = None,
) -> str:
    """
    Ponto de entrada da Fase 3.
    Recebe o DataFrame higienizado (ou o lê do CSV) e gera a planilha final.
    Retorna o caminho do arquivo gerado.
    """
    if df_higienizado is None:
        caminho_csv = BASE_DIR / "dados_brutos" / "comunicacoes_higienizadas.csv"
        if not caminho_csv.exists():
            raise FileNotFoundError(
                f"Arquivo '{caminho_csv}' não encontrado. Execute a Fase 2 primeiro."
            )
        df_higienizado = pd.read_csv(caminho_csv, dtype=str).fillna("")

    alertas_f3: list[dict] = []

    logger.info(f"Fase 3: classificando {len(df_higienizado)} comunicação(ões)…")

    # 1. Aplica classificação linha a linha
    resultados = df_higienizado["texto_higienizado"].apply(classificar_comunicacao)

    df_higienizado["temas"]      = resultados.apply(lambda r: r["temas"])
    df_higienizado["score"]      = resultados.apply(lambda r: r["score"])
    df_higienizado["bonus_desc"] = resultados.apply(lambda r: r["bonus_desc"])
    df_higienizado["evidencias"] = resultados.apply(lambda r: r["evidencias"])

    # 2. Valida scores negativos (não deveria ocorrer, mas tratamos)
    negativos = df_higienizado[df_higienizado["score"] < 0]
    if not negativos.empty:
        alertas_f3.append({
            "tipo":      "score_negativo",
            "descricao": f"{len(negativos)} registro(s) com score negativo — verifique os pesos.",
        })

    # 3. Consolida por processo
    df_resumo = consolidar_por_processo(df_higienizado)

    # 4. Junta todos os alertas
    todos_alertas = (alertas_anteriores or []) + alertas_f3

    # 5. Gera planilha
    gerar_planilha(df_higienizado, df_resumo, todos_alertas, ARQUIVO_SAIDA)

    logger.info(
        f"Fase 3 concluída: {len(df_higienizado)} comunicação(ões) | "
        f"{df_resumo['score_maximo'].max() if not df_resumo.empty else 0} score máximo."
    )

    return ARQUIVO_SAIDA


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    executar()
